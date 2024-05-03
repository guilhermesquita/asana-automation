import requests
import openpyxl
import os
from dotenv import load_dotenv

offset = ''
data_task = []

load_dotenv()

asana_api_key = os.getenv("ASANA_API_KEY")
asana_api_url = os.getenv("ASANA_API_URL")
arquivo_xlsx = os.getenv("PLANILHA_XLSX")

url_base = asana_api_url + 'sections/1201422389995890/tasks?opt_fields=&limit=10&{offset}'.format(
    offset=offset)
headers = {
    'Authorization': 'Bearer ' + asana_api_key,
    'Content-Type': 'application/json'
}


def get_task_data(url, headers):
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        print('Falha na requisição. Código de status:', response.status_code)
        return None

while True:
    data = get_task_data(url_base, headers)
    if not data:
        break

    tasks = data.get('data')
    if tasks:
        for task in tasks:
            gid = task['gid']
            url_task = "https://app.asana.com/api/1.0/tasks/{gid}".format(gid=gid)
            task_data = get_task_data(url_task, headers)
            if not task_data:
                continue
            data_json = task_data.get('data')

            indice_inicio_telefone = data_json['notes'].find('Telefone:')

            if indice_inicio_telefone != -1:
                indice_fim_telefone = data_json['notes'].find('\n', indice_inicio_telefone)                
                if indice_fim_telefone != -1:
                    number_telephone = data_json['notes'][indice_inicio_telefone+len('Telefone:'):indice_fim_telefone].strip()

            if not data_json:
                continue

            url_story = "https://app.asana.com/api/1.0/tasks/{gid}/stories".format(gid=gid)
            story_data = get_task_data(url_story, headers)
            if not story_data:
                continue

            data_story = story_data.get('data')

            processnumber = None

            for story in data_story:
                if processnumber is not None:
                    break
    
                text = story['text'].lower()
                keywords = ['processo:', 'processo ', 'número do processo:', 'número do processo ', 'processo nº:', 'processo nº ']
    
                for keyword in keywords:
                    if keyword in text:
                        processnumber = text.split(keyword)[1].split()[0]
                        break

            atuations = ['Voo cancelado', 'Bagagem extraviada']

            involved = objective = None
            for atuation in atuations:
                if len(data_json['tags']) > 2:
                    if atuation.lower() == data_json['tags'][1]['name'].lower():
                        objective = data_json['tags'][1]['name']
                        involved = data_json['tags'][2]['name']
                if len(data_json['tags']) > 2: 
                    if atuation.lower() == data_json['tags'][2]['name'].lower():
                        objective = data_json['tags'][2]['name']
                        involved = data_json['tags'][1]['name']
                if len(data_json['tags']) == 2:
                    if atuation.lower() == data_json['tags'][1]['name'].lower():
                        objective = data_json['tags'][1]['name']
                    else:
                        involved = data_json['tags'][1]['name']

            data_task.append({'name': data_json['name'], 'number': processnumber, 'involved': involved,
                              'objective': objective,'telephone': number_telephone})
            print(data_task)

        if data.get('next_page'):
            offset = 'offset=' + data['next_page']['offset']
            url_base = 'https://app.asana.com/api/1.0/sections/1201422389995890/tasks?opt_fields=&limit=100&{offset}'.format(
                offset=offset)
        else:
            break

    else:
        print('Não há mais tarefas.')
        break

print(data_task)
def preencher_planilha(data_task, planilha_excel):
    wb = openpyxl.load_workbook(planilha_excel)
    ws = wb['Processos  - Preenchimento']
    ws2 = wb['Contatos - Preenchimento']

    for row, task_data in enumerate(data_task, start=3):  
        ws.cell(row=row, column=2, value=task_data.get('name', ''))  
        ws.cell(row=row, column=5, value=task_data.get('number', ''))  
        ws.cell(row=row, column=6, value=0)
        ws.cell(row=row, column=7, value='indenizatória')
        ws.cell(row=row, column=8, value=1)
        ws.cell(row=row, column=15, value=task_data.get('name', ''))  
        ws.cell(row=row, column=17, value='Reclamante')  
        ws.cell(row=row, column=18, value=task_data.get('involved', ''))  
        ws.cell(row=row, column=20, value=task_data.get('objective', ''))  
        ws.cell(row=row, column=23, value=task_data.get('objective', ''))  
        ws.cell(row=row, column=30, value='contato@barbosaluan.com')   

    for row, task_data in enumerate(data_task, start=3):  
        ws2.cell(row=row, column=1, value=task_data.get('name', ''))  
        ws2.cell(row=row, column=5, value=task_data.get('telephone', ''))    


    wb.save(planilha_excel)

planilha_existente = 'arquivo_xlsx'
preencher_planilha(data_task, planilha_existente)
print("Informações já exportadas")